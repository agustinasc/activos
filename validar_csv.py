"""
Validador de CSV de nómina - AG.A.P. Catamarca
Genera un Excel con resumen y detalle de errores.
"""
import pandas as pd
import re
import sys
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# COLUMNAS DEL CSV
# ---------------------------------------------------------------------------
COLUMNAS = [
    'cuil','codobra','canthijo','conyuge','adherente','rebpromo',
    'zonageo','actividad','porcreduc','remutotal','remuimpo','porceadi',
    'aportvolun','exessocial','asigfliar','exeosocial','remuimpoos',
    'impoadicos','aportejubi','aportepatr','apafilos','appatros',
    'apeynom','domicilio','fechaing','fechabaja','conceptono','codigo',
    'declara','periodo','codorigen','nroagente','ubicacion','coseguro',
    'codestado','saf','d_svco','d_ss','d_svca','d_sc','remusueldo',
    'remusac','hsextras','dias_traba','canasta_fl','tipoliq','funcionari',
    'remu_doc','act_doc','ap_dcto788','aporte_dif','cant_hsext',
    'fecha_nac','ad_caja_co','aporte_caj','contrib_ca','tiene_deud',
    'deuda_caja','trasp_osep','tiene_fdo_','sexo','premios','vacaciones',
    'remuimpoto','por_retiro','rem_retiro','apo_vol','tras_vol',
    'cos_pro','cos_alt','cos_int','claseycar'
]

# ---------------------------------------------------------------------------
# PARÁMETROS Y TOPES - Actualizar cada período
# ---------------------------------------------------------------------------

# ANSES RS 110/2026
ANSES_BASE_MIN    = 132420.94
ANSES_BASE_MAX    = 4303619.01
ANSES_APORTE_MIN  = 14566.3
ANSES_APORTE_MAX  = 473398.09
ANSES_CONTRIB_MIN = 14239.22

# Caja Complementaria (Municipios / Poder Ejecutivo / Poder Judicial)
CAJA_BASE_MIN    = 363000.00
CAJA_APORTE_MIN  = 10890.00
CAJA_APORTE_MAX  = 235056.6   # Municipios/Ejecutivo (Judicial: 387532.16)
CAJA_PCT_APORTE  = 0.03
CAJA_PCT_CONTRIB = 0.03

# Obra Social OSEP Dcto 142/25
OSEP_APORTE_MIN  = 34828.32    # apafilos mínimo
OSEP_CONTRIB_MIN = 69656.64   # appatros mínimo
OSEP_TRASP_MIN   = 3869.81     # trasp_osep mínimo
OSEP_PCT_APORTE  = 0.045
OSEP_PCT_CONTRIB = 0.09
OSEP_PCT_TRASP   = 0.005       # 0.5% de remuimpoos
OSEP_BASE_MIN    = 773962.74
APO_VOL_MIN_1    = round(OSEP_BASE_MIN * 0.025, 2)   # $19.349,07 (1 adherente)
APO_VOL_MIN_2    = round(OSEP_BASE_MIN * 0.05,  2)   # $38.698,14 (2+ adherentes)
TRAS_VOL_MIN     = round(OSEP_BASE_MIN * 0.005, 2)   # $3.869,81

# Seguros fijos
D_SVCO_FIJO = 5000.00
D_SS_FIJO   = 4400.00

# Seguridad Social
SS_PCT_APORTE       = 0.11
SS_PCT_CONTRIB      = 0.10753
SS_PCT_MAGISTRADO   = 0.18     # aportejubi magistrado (act_doc='01')

# Aporte voluntario
APO_VOL_PCT_1 = 0.025
APO_VOL_PCT_2 = 0.05
TRAS_VOL_PCT  = 0.005


# Actividades permitidas
ACTIVIDADES_PERMITIDAS = [
    "1106","2106","1206","2206","1306","2306",
    "1406","2406","9999","1506","2506",
    "0106","0206","0306","0406","01"
]

ACT_DOC_PERMITIDOS  = ["75","76","77","78","79","00","01"]
CODOBRA_PERMITIDOS  = ["009960","001102"]
EDAD_MIN_INGRESO    = 18
TOLERANCIA          = 0.02

# ---------------------------------------------------------------------------
# DESCRIPCIONES Y SEVERIDAD DE ERRORES
# ---------------------------------------------------------------------------
ERRORES = {
    1:  ('CRÍTICO',      'apeynom vacío'),
    2:  ('CRÍTICO',      'fecha_nac vacía'),
    3:  ('ADVERTENCIA',  'actividad no permitida'),
    4:  ('CRÍTICO',      'nroagente vacío'),
    5:  ('CRÍTICO',      'CUIL inválido'),
    6:  ('ADVERTENCIA',  'saf vacío'),
    7:  ('ADVERTENCIA',  'ubicacion vacía'),
    8:  ('ADVERTENCIA',  'codobra inválido (debe ser 009960 o 001102)'),
    9:  ('CRÍTICO',      'remutotal formato inválido'),
    10: ('CRÍTICO',      'remuimpo formato inválido'),
    11: ('ADVERTENCIA',  'remuimpoos formato inválido'),
    13: ('CRÍTICO',      'periodo inválido o no coincide con el período declarado'),
    17: ('CRÍTICO',      'CUIL duplicado'),
    18: ('ADVERTENCIA',  'aportejubi menor al mínimo ANSES'),
    19: ('ADVERTENCIA',  'aporte_caj menor al mínimo caja complementaria'),
    22: ('ADVERTENCIA',  'zonageo inválida (debe ser 08 o 09)'),
    23: ('ADVERTENCIA',  'act_doc inválido para remuneración docente'),
    24: ('ADVERTENCIA',  'aporte_caj no coincide con el cálculo esperado'),
    25: ('ADVERTENCIA',  'contrib_ca no coincide con el cálculo esperado'),
    26: ('ADVERTENCIA',  'deuda_caja no coincide con el cálculo esperado'),
    30: ('ADVERTENCIA',  'remuimpoos debe ser igual a remutotal'),
    31: ('ADVERTENCIA',  'nroagente duplicado'),
    32: ('ADVERTENCIA',  'domicilio vacío'),
    35: ('ADVERTENCIA',  'codorigen debe ser 0'),
    36: ('ADVERTENCIA',  'remusac debe tener valor en junio o diciembre (SAC)'),
    38: ('CRÍTICO',      'dias_traba debe estar entre 1 y 30'),
    39: ('ADVERTENCIA',  'ad_caja_co debe ser F o T'),
    40: ('ADVERTENCIA',  'tiene_deud debe ser F o T'),
    41: ('ADVERTENCIA',  'trasp_osep no coincide con remuimpoos * 0.5%'),
    42: ('CRÍTICO',      'sexo debe ser F, M o X'),
    43: ('ADVERTENCIA',  'premios formato inválido'),
    44: ('ADVERTENCIA',  'vacaciones formato inválido'),
    45: ('ADVERTENCIA',  'remutotal no coincide con el cálculo esperado'),
    46: ('ADVERTENCIA',  'remuimpo no coincide con el cálculo esperado'),
    47: ('ADVERTENCIA',  'remuimpo mayor que remutotal'),
    48: ('ADVERTENCIA',  'codigo de organismo no coincide'),
    49: ('ADVERTENCIA',  'tipoliq debe ser 1'),
    50: ('ADVERTENCIA',  'apafilos no coincide con remutotal * 4.5%'),
    51: ('ADVERTENCIA',  'appatros no coincide con remutotal * 9%'),
    52: ('ADVERTENCIA',  'aportepatr no coincide con remuimpo * 10.17%'),
    53: ('ADVERTENCIA',  'aportejubi no coincide con el cálculo esperado'),
    54: ('ADVERTENCIA',  f'd_svco debe ser {D_SVCO_FIJO:.2f}'),
    55: ('ADVERTENCIA',  f'd_ss debe ser {D_SS_FIJO:.2f}'),
    60: ('ADVERTENCIA',  'fechaing vacía'),
    61: ('ADVERTENCIA',  'fechaing formato inválido (debe ser YYYYMMDD)'),
    62: ('ADVERTENCIA',  'fecha_nac formato inválido (debe ser YYYYMMDD)'),
    63: ('ADVERTENCIA',  f'agente menor de {EDAD_MIN_INGRESO} años al momento de ingreso'),
    64: ('ADVERTENCIA',  'fechaing anterior a fecha_nac'),
    65: ('ADVERTENCIA',  f'apafilos menor al mínimo OSEP (${OSEP_APORTE_MIN:,.2f})'),
    66: ('ADVERTENCIA',  f'appatros menor al mínimo OSEP (${OSEP_CONTRIB_MIN:,.2f})'),
    67: ('ADVERTENCIA',  f'trasp_osep menor al mínimo OSEP (${OSEP_TRASP_MIN:,.2f})'),
    68: ('ADVERTENCIA',  'apo_vol no es múltiplo de 2.5% de remuimpoos'),
    69: ('ADVERTENCIA',  'tras_vol vacío cuando apo_vol > 0'),
    70: ('ADVERTENCIA',  'tras_vol no coincide con remuimpoos * 0.5%'),
    71: ('ADVERTENCIA',  f'aportejubi supera el tope máximo ANSES (${ANSES_APORTE_MAX:,.2f})'),
    72: ('ADVERTENCIA',  'd_svco vacío o cero'),
    73: ('ADVERTENCIA',  'd_ss vacío o cero'),
    74: ('ADVERTENCIA',  'trasp_osep vacío o cero'),
}

# ---------------------------------------------------------------------------
# FUNCIONES AUXILIARES
# ---------------------------------------------------------------------------
def v(fila, campo):
    val = fila.get(campo, '')
    return str(val).strip() if pd.notna(val) else ''

def n(fila, campo):
    s = v(fila, campo)
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0

def es_num(s):
    return bool(re.match(r'^[0-9]{1,8}(\.[0-9]{1,2})?$', s.strip())) if s.strip() else False

def igual(a, b, tol=TOLERANCIA):
    return abs(round(float(a), 2) - round(float(b), 2)) <= tol

def fecha_valida(s):
    s = str(s).strip()
    if len(s) != 8 or not s.isdigit():
        return False
    try:
        datetime.strptime(s, '%Y%m%d')
        return True
    except ValueError:
        return False

def validar_cuil(cuil):
    cuil = str(cuil).strip()
    if not re.match(r'^\d{11}$', cuil):
        return False
    mult = [5,4,3,2,7,6,5,4,3,2]
    suma = sum(int(cuil[i]) * mult[i] for i in range(10))
    mod = 11 - (suma % 11)
    if mod == 11: mod = 0
    if mod == 10: mod = 9
    return mod == int(cuil[10])

def normalizar_periodo(p):
    """Convierte AAAAMM a MMAAAA si es necesario."""
    p = str(p).strip()
    if len(p) == 6 and p.isdigit() and int(p[:2]) > 12:
        return p[4:6] + p[0:4]
    return p

# ---------------------------------------------------------------------------
# VALIDADOR PRINCIPAL
# ---------------------------------------------------------------------------
def validar_csv(csv_path):
    df = pd.read_csv(csv_path, header=None, encoding='latin1', dtype=str)
    df.columns = COLUMNAS[:len(df.columns)]

    # Detectar y normalizar período del archivo
    periodos = df['periodo'].dropna().unique()
    periodo_declarado = normalizar_periodo(periodos[0]) if len(periodos) > 0 else ''

    errores      = []
    cuils_vistos = []
    agentes_vistos = []

    for idx, fila in df.iterrows():
        nro_linea = idx + 1
        cuil      = v(fila, 'cuil')
        agente    = v(fila, 'nroagente')
        apeynom   = v(fila, 'apeynom')
        es_magistrado = v(fila, 'act_doc').strip() == '01'

        def agregar(cod):
            sev, desc = ERRORES.get(cod, ('ADVERTENCIA', f'Error {cod}'))
            errores.append({
                'Línea':             nro_linea,
                'CUIL':              cuil,
                'Agente':            agente,
                'Apellido y Nombre': apeynom,
                'Magistrado':        'Sí' if es_magistrado else 'No',
                'Código':            cod,
                'Severidad':         sev,
                'Descripción':       desc,
            })

        # --- Campos obligatorios vacíos ---
        if not v(fila, 'apeynom'):   agregar(1)
        if not v(fila, 'fecha_nac'): agregar(2)
        if not v(fila, 'saf'):       agregar(6)
        if not v(fila, 'ubicacion'): agregar(7)
        if not v(fila, 'domicilio'): agregar(32)
        if not v(fila, 'nroagente'): agregar(4)
        if not v(fila, 'fechaing'):  agregar(60)

        # --- CUIL ---
        if not validar_cuil(cuil):
            agregar(5)
        else:
            if cuil in cuils_vistos:
                agregar(17)
            else:
                cuils_vistos.append(cuil)

        # --- nroagente duplicado ---
        if agente:
            if agente in agentes_vistos:
                agregar(31)
            else:
                agentes_vistos.append(agente)

        # --- Códigos fijos ---
        if v(fila, 'codobra') not in CODOBRA_PERMITIDOS: agregar(8)
        if v(fila, 'actividad') not in ACTIVIDADES_PERMITIDAS: agregar(3)
        if v(fila, 'zonageo') not in ['08', '09']:        agregar(22)
        if v(fila, 'codorigen') != '0':                   agregar(35)
        if v(fila, 'tipoliq') != '1':                     agregar(49)
        if v(fila, 'sexo') not in ['F', 'M', 'X']:        agregar(42)
        if v(fila, 'ad_caja_co') not in ['F', 'T']:       agregar(39)
        if v(fila, 'tiene_deud') not in ['F', 'T']:       agregar(40)

        # --- dias_traba ---
        dias = n(fila, 'dias_traba')
        if not (1 <= dias <= 30):
            agregar(38)

        # --- Período ---
        per = normalizar_periodo(v(fila, 'periodo'))
        per_ok = False
        if len(per) == 6 and per.isdigit():
            mes  = per[:2]
            anio = per[2:]
            try:
                datetime(int(anio), int(mes), 1)
                per_ok = True
                if per != periodo_declarado:
                    agregar(13)
                else:
                    if int(mes) in [6, 12] and n(fila, 'remusac') == 0:
                        agregar(36)
            except ValueError:
                pass
        if not per_ok:
            agregar(13)

        # --- Fechas ---
        fi = v(fila, 'fechaing')
        fn = v(fila, 'fecha_nac')
        fi_ok = fecha_valida(fi)
        fn_ok = fecha_valida(fn)

        if fi and not fi_ok: agregar(61)
        if fn and not fn_ok: agregar(62)

        if fi_ok and fn_ok:
            dt_fi = datetime.strptime(fi, '%Y%m%d')
            dt_fn = datetime.strptime(fn, '%Y%m%d')
            if dt_fi < dt_fn:
                agregar(64)
            else:
                edad = (dt_fi - dt_fn).days // 365
                if edad < EDAD_MIN_INGRESO:
                    agregar(63)

        # --- Formatos numéricos ---
        if not es_num(v(fila, 'remutotal')):  agregar(9)
        if not es_num(v(fila, 'remuimpo')):   agregar(10)
        if not es_num(v(fila, 'remuimpoos')): agregar(11)
        if not es_num(v(fila, 'premios')):    agregar(43)
        if not es_num(v(fila, 'vacaciones')): agregar(44)

        # --- Valores numéricos ---
        remutotal  = n(fila, 'remutotal')
        remuimpo   = n(fila, 'remuimpo')
        remuimpoos = n(fila, 'remuimpoos')
        remusueldo = n(fila, 'remusueldo')
        remusac    = n(fila, 'remusac')
        hsextras   = n(fila, 'hsextras')
        conceptono = n(fila, 'conceptono')
        premios    = n(fila, 'premios')
        vacaciones = n(fila, 'vacaciones')
        canasta    = n(fila, 'canasta_fl')

        # --- Cálculos de remuneraciones ---
        calc_remuimpo = round(remusueldo + remusac + hsextras, 2)
        if not igual(remuimpo, calc_remuimpo): agregar(46)
        elif remuimpo > remutotal:             agregar(47)

        calc_remutotal = round(remuimpo + conceptono + premios + vacaciones + canasta, 2)
        if not igual(remutotal, calc_remutotal): agregar(45)

        if not igual(remuimpoos, remutotal): agregar(30)

        # --- aportejubi ---
        aportejubi = n(fila, 'aportejubi')
        if es_magistrado:
            # Magistrado: 18% de remuimpo, sin tope máximo
            calc_aj = round(remuimpo * SS_PCT_MAGISTRADO, 2)
            if not igual(aportejubi, calc_aj):
                agregar(53)
        else:
            # No magistrado: 11% hasta el tope ANSES
            if remuimpo <= ANSES_BASE_MAX:
                calc_aj = round(remuimpo * SS_PCT_APORTE, 2)
                # Aceptar también si se aplicó el tope máximo directamente
                if not igual(aportejubi, calc_aj) and not igual(aportejubi, ANSES_APORTE_MAX):
                    agregar(53)
            else:
                # Remuimpo supera el tope: aportejubi debe ser <= ANSES_APORTE_MAX
                pass
            if aportejubi > ANSES_APORTE_MAX + TOLERANCIA and not es_magistrado:
                agregar(71)
            if dias == 30 and aportejubi < ANSES_APORTE_MIN - TOLERANCIA:
                agregar(18)

        # --- aportepatr: 10.17% de remuimpo, sin tope ---
        aportepatr = n(fila, 'aportepatr')
        if not igual(aportepatr, round(remuimpo * SS_PCT_CONTRIB, 2)):
            agregar(52)

        # --- apafilos: 4.5% de remutotal, mínimo OSEP ---
        apafilos = n(fila, 'apafilos')
        if apafilos < OSEP_APORTE_MIN - TOLERANCIA:
            agregar(65)   # por debajo del mínimo -> advertencia
        elif not igual(apafilos, round(remutotal * OSEP_PCT_APORTE, 2)) and              not igual(apafilos, OSEP_APORTE_MIN):
            agregar(50)   # no coincide con % ni con mínimo -> advertencia

        # --- appatros: 9% de remutotal, mínimo OSEP ---
        appatros = n(fila, 'appatros')
        if appatros < OSEP_CONTRIB_MIN - TOLERANCIA:
            agregar(66)   # por debajo del mínimo -> advertencia
        elif not igual(appatros, round(remutotal * OSEP_PCT_CONTRIB, 2)) and              not igual(appatros, OSEP_CONTRIB_MIN):
            agregar(51)   # no coincide con % ni con mínimo -> advertencia

        # --- trasp_osep: 0.5% de remuimpoos, solo para OSEP (codobra=009960) ---
        trasp_osep = n(fila, 'trasp_osep')
        if v(fila, 'codobra') == '009960':
            if trasp_osep == 0:
                agregar(74)
            else:
                # Si el valor >= mínimo es válido (puede ser el mínimo aplicado)
                if trasp_osep < OSEP_TRASP_MIN - TOLERANCIA:
                    agregar(67)
                elif not igual(trasp_osep, round(remuimpoos * OSEP_PCT_TRASP, 2)) and                      not igual(trasp_osep, OSEP_TRASP_MIN):
                    agregar(41)

        # --- Seguros fijos ---
        if not v(fila, 'd_svco') or n(fila, 'd_svco') == 0: agregar(72)
        elif not igual(n(fila, 'd_svco'), D_SVCO_FIJO):     agregar(54)

        if not v(fila, 'd_ss') or n(fila, 'd_ss') == 0: agregar(73)
        elif not igual(n(fila, 'd_ss'), D_SS_FIJO):      agregar(55)

        # --- Caja complementaria ---
        if v(fila, 'ad_caja_co') == 'T':
            aporte_caj = n(fila, 'aporte_caj')
            contrib_ca = n(fila, 'contrib_ca')
            if not igual(aporte_caj, round(remutotal * CAJA_PCT_APORTE, 2)):
                agregar(24)
            if dias == 30 and aporte_caj < CAJA_APORTE_MIN - TOLERANCIA:
                agregar(19)
            if not igual(contrib_ca, round(remutotal * CAJA_PCT_CONTRIB, 2)):
                agregar(25)

        if v(fila, 'tiene_deud') == 'T' and CAJA_PCT_APORTE > 0:
            deuda_caja = n(fila, 'deuda_caja')
            if deuda_caja == 0:
                agregar(26)

        # --- Remuneración docente ---
        if n(fila, 'remu_doc') > 0:
            if v(fila, 'act_doc').strip() not in ACT_DOC_PERMITIDOS:
                agregar(23)

        # --- codigo organismo ---
        if not v(fila, 'codigo'):
            agregar(48)

        # --- apo_vol / tras_vol ---
        apo_vol  = n(fila, 'apo_vol')
        tras_vol = n(fila, 'tras_vol')
        try:
            cant_adh = int(v(fila, 'adherente'))
        except ValueError:
            cant_adh = 0

        if apo_vol > 0:
            pct      = APO_VOL_PCT_2 if cant_adh >= 2 else APO_VOL_PCT_1
            vol_min  = APO_VOL_MIN_2  if cant_adh >= 2 else APO_VOL_MIN_1
            calc_vol = round(remuimpoos * pct, 2)
            # Válido solo si coincide con el % o está aportando exactamente el mínimo
            if not igual(apo_vol, calc_vol) and not igual(apo_vol, vol_min):
                if apo_vol < vol_min - TOLERANCIA:
                    agregar(68)   # por debajo del mínimo
                else:
                    agregar(68)   # no coincide con % ni con mínimo

            if tras_vol == 0:
                agregar(69)
            else:
                calc_tv = round(remuimpoos * TRAS_VOL_PCT, 2)
                # Válido solo si coincide con el % o está aportando exactamente el mínimo
                if not igual(tras_vol, calc_tv) and not igual(tras_vol, TRAS_VOL_MIN):
                    if tras_vol < TRAS_VOL_MIN - TOLERANCIA:
                        agregar(70)   # por debajo del mínimo
                    else:
                        agregar(70)   # no coincide con % ni con mínimo

    return errores, periodo_declarado, len(df)


# ---------------------------------------------------------------------------
# GENERAR EXCEL DE REPORTE
# ---------------------------------------------------------------------------
def generar_excel(errores, csv_path, periodo, total_registros, output_path):
    wb = Workbook()

    hdr_font   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill_r = PatternFill('solid', start_color='C00000')
    hdr_fill_a = PatternFill('solid', start_color='ED7D31')
    hdr_fill_b = PatternFill('solid', start_color='1F4E79')
    alt_fill_r = PatternFill('solid', start_color='FCE4D6')
    alt_fill_a = PatternFill('solid', start_color='FFF2CC')
    alt_fill_w = PatternFill('solid', start_color='FFFFFF')
    thin       = Side(style='thin', color='CCCCCC')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal='center', vertical='center')
    left       = Alignment(horizontal='left',   vertical='center')

    criticos     = [e for e in errores if e['Severidad'] == 'CRÍTICO']
    advertencias = [e for e in errores if e['Severidad'] == 'ADVERTENCIA']

    # --- Hoja Resumen ---
    ws = wb.active
    ws.title = 'Resumen'
    ws.merge_cells('A1:D1')
    ws['A1'] = f'Validación CSV — {os.path.basename(csv_path)}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')

    estado     = 'CON ERRORES CRÍTICOS' if criticos else ('CON ADVERTENCIAS' if advertencias else 'OK ✓')
    color_est  = 'C00000' if criticos else ('ED7D31' if advertencias else '375623')

    resumen = [
        ('Archivo',          os.path.basename(csv_path)),
        ('Período',          periodo),
        ('Total registros',  total_registros),
        ('Errores CRÍTICOS', len(criticos)),
        ('Advertencias',     len(advertencias)),
        ('Total errores',    len(errores)),
        ('Estado',           estado),
    ]
    for i, (label, valor) in enumerate(resumen, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(name='Arial', bold=True, size=10)
        c = ws.cell(row=i, column=2, value=valor)
        c.font = Font(name='Arial', size=10,
                      color=color_est if label == 'Estado' else '000000',
                      bold=(label == 'Estado'))

    ws.cell(row=12, column=1, value='Detalle por tipo de error').font = Font(
        name='Arial', bold=True, size=11, color='1F4E79')

    for j, h in enumerate(['Cód','Descripción','Severidad','Cantidad'], 1):
        c = ws.cell(row=13, column=j, value=h)
        c.font = hdr_font; c.fill = hdr_fill_b
        c.alignment = center; c.border = border

    conteo = {}
    for e in errores:
        key = (e['Código'], e['Descripción'], e['Severidad'])
        conteo[key] = conteo.get(key, 0) + 1

    for ri, ((cod, desc, sev), cnt) in enumerate(sorted(conteo.items()), start=14):
        fill = alt_fill_r if sev == 'CRÍTICO' else alt_fill_a
        for j, val in enumerate([cod, desc, sev, cnt], 1):
            c = ws.cell(row=ri, column=j, value=val)
            c.font = Font(name='Arial', size=9); c.fill = fill
            c.border = border
            c.alignment = center if j in [1,3,4] else left

    for col, ancho in zip('ABCD', [6, 52, 16, 10]):
        ws.column_dimensions[col].width = ancho

    # --- Función para crear hojas de errores ---
    COLS = ['Línea','CUIL','Agente','Apellido y Nombre','Magistrado','Código','Severidad','Descripción']
    ANCHOS = [7, 13, 12, 35, 10, 7, 14, 55]

    def crear_hoja(nombre, lista, fill_hdr, fill_par):
        ws2 = wb.create_sheet(nombre)
        for j, h in enumerate(COLS, 1):
            c = ws2.cell(row=1, column=j, value=h)
            c.font = hdr_font; c.fill = fill_hdr
            c.alignment = center; c.border = border
        for i, err in enumerate(lista, start=2):
            fill = fill_par if i % 2 == 0 else alt_fill_w
            for j, col in enumerate(COLS, 1):
                c = ws2.cell(row=i, column=j, value=err.get(col, ''))
                c.font = Font(name='Arial', size=9); c.fill = fill
                c.border = border
                c.alignment = center if j in [1,2,3,5,6,7] else left
        for j, ancho in enumerate(ANCHOS, 1):
            ws2.column_dimensions[get_column_letter(j)].width = ancho
        ws2.freeze_panes = 'A2'
        ws2.row_dimensions[1].height = 20

    crear_hoja('Críticos',        criticos,     hdr_fill_r, alt_fill_r)
    crear_hoja('Advertencias',    advertencias, hdr_fill_a, alt_fill_a)
    crear_hoja('Todos los errores', errores,   hdr_fill_b, alt_fill_a)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main(csv_path, output_path=None):
    if output_path is None:
        base = os.path.splitext(csv_path)[0]
        output_path = base + '_validacion.xlsx'

    print(f"Validando: {csv_path}")
    errores, periodo, total = validar_csv(csv_path)

    criticos     = sum(1 for e in errores if e['Severidad'] == 'CRÍTICO')
    advertencias = sum(1 for e in errores if e['Severidad'] == 'ADVERTENCIA')

    generar_excel(errores, csv_path, periodo, total, output_path)

    print(f"Período detectado : {periodo}")
    print(f"Registros         : {total}")
    print(f"Errores CRÍTICOS  : {criticos}")
    print(f"Advertencias      : {advertencias}")
    print(f"Total errores     : {len(errores)}")
    print(f"Reporte generado  : {output_path}")

    if criticos > 0:
        print("\n  ⚠ Hay errores CRÍTICOS — revisar antes de generar los TXT.")
    elif advertencias > 0:
        print("\n  ⚠ Hay advertencias — revisar antes de enviar.")
    else:
        print("\n  ✓ Sin errores detectados.")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python validar_csv.py <archivo.csv> [reporte.xlsx]")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
