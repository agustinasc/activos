import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os
import re

# Columnas en el orden exacto del archivo CSV
COLUMNAS = [
    'cuil', 'codobra', 'canthijo', 'conyuge', 'adherente', 'rebpromo',
    'zonageo', 'actividad', 'porcreduc', 'remutotal', 'remuimpo', 'porceadi',
    'aportvolun', 'exessocial', 'asigfliar', 'exeosocial', 'remuimpoos',
    'impoadicos', 'aportejubi', 'aportepatr', 'apafilos', 'appatros',
    'apeynom', 'domicilio', 'fechaing', 'fechabaja', 'conceptono', 'codigo',
    'declara', 'periodo', 'codorigen', 'nroagente', 'ubicacion', 'coseguro',
    'codestado', 'saf', 'd_svco', 'd_ss', 'd_svca', 'd_sc', 'remusueldo',
    'remusac', 'hsextras', 'dias_traba', 'canasta_fl', 'tipoliq', 'funcionari',
    'remu_doc', 'act_doc', 'ap_dcto788', 'aporte_dif', 'cant_hsext',
    'fecha_nac', 'ad_caja_co', 'aporte_caj', 'contrib_ca', 'tiene_deud',
    'deuda_caja', 'trasp_osep', 'tiene_fdo_', 'sexo', 'premios', 'vacaciones',
    'remuimpoto', 'por_retiro', 'rem_retiro', 'apo_vol', 'tras_vol',
    'cos_pro', 'cos_alt', 'cos_int', 'claseycar'
]

# Columnas numéricas con importes (para formatear con 2 decimales)
COLS_IMPORTE = {
    'remutotal', 'remuimpo', 'aportvolun', 'exessocial', 'asigfliar',
    'exeosocial', 'remuimpoos', 'impoadicos', 'aportejubi', 'aportepatr',
    'apafilos', 'appatros', 'conceptono', 'remusueldo', 'remusac',
    'canasta_fl', 'remu_doc', 'ap_dcto788', 'aporte_dif', 'aporte_caj',
    'contrib_ca', 'deuda_caja', 'trasp_osep', 'premios', 'vacaciones',
    'remuimpoto', 'rem_retiro', 'apo_vol', 'tras_vol', 'cos_pro',
    'cos_alt', 'cos_int'
}

# Columnas de fecha (formato YYYYMMDD -> DD/MM/YYYY)
COLS_FECHA = {'fechaing', 'fechabaja', 'fecha_nac'}


def extraer_periodo_del_nombre(nombre_archivo):
    """Extrae el periodo MMAAAA del nombre del archivo, ej: SAN_FERNANDO_032026 -> 032026"""
    match = re.search(r'_(\d{6})(?:\.csv)?$', nombre_archivo, re.IGNORECASE)
    if match:
        return match.group(1)
    return ""


def formatear_fecha(valor):
    """Convierte YYYYMMDD o fecha a DD/MM/YYYY"""
    try:
        s = str(int(float(valor)))
        if len(s) == 8:
            return f"{s[6:8]}/{s[4:6]}/{s[0:4]}"
    except (ValueError, TypeError):
        pass
    return valor


def main(csv_path, output_path=None):
    nombre = os.path.basename(csv_path)
    periodo = extraer_periodo_del_nombre(nombre)

    if output_path is None:
        base = os.path.splitext(nombre)[0]
        output_path = os.path.join(os.path.dirname(csv_path), f"{base}.xlsx")

    # Leer CSV sin header
    df = pd.read_csv(csv_path, header=None, encoding='latin1', dtype=str)
    df.columns = COLUMNAS[:len(df.columns)]

    # Formatear fechas
    for col in COLS_FECHA:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: formatear_fecha(x) if pd.notna(x) and str(x).strip() not in ('', 'nan') else '')

    # Escribir con openpyxl para aplicar formato
    df.to_excel(output_path, index=False, sheet_name='Datos')

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = 'Datos'

    # Estilos header
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Formato filas de datos
    data_font = Font(name='Arial', size=9)
    data_align_left = Alignment(horizontal='left', vertical='center')
    data_align_right = Alignment(horizontal='right', vertical='center')
    data_align_center = Alignment(horizontal='center', vertical='center')
    fill_alt = PatternFill('solid', start_color='EBF3FB')

    num_format_importe = '#,##0.00'
    num_format_entero = '#,##0'

    # Aplicar header
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Aplicar formato a datos
    for row_idx in range(2, ws.max_row + 1):
        fill = fill_alt if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = border

            if fill:
                cell.fill = fill

            if col_name in COLS_IMPORTE:
                cell.alignment = data_align_right
                try:
                    val = float(str(cell.value).replace(',', '.')) if cell.value not in (None, '', 'nan') else None
                    if val is not None:
                        cell.value = val
                        cell.number_format = num_format_importe
                except (ValueError, TypeError):
                    pass
            elif col_name in ('cuil', 'nroagente', 'ubicacion', 'codobra', 'nroagente'):
                cell.alignment = data_align_center
            elif col_name in COLS_FECHA:
                cell.alignment = data_align_center
            else:
                cell.alignment = data_align_left

    # Ajustar ancho de columnas
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        if col_name in ('apeynom', 'domicilio', 'ubicacion'):
            ws.column_dimensions[col_letter].width = 28
        elif col_name in ('cuil',):
            ws.column_dimensions[col_letter].width = 14
        elif col_name in COLS_IMPORTE:
            ws.column_dimensions[col_letter].width = 14
        elif col_name in COLS_FECHA:
            ws.column_dimensions[col_letter].width = 12
        else:
            ws.column_dimensions[col_letter].width = 10

    # Fila de altura header
    ws.row_dimensions[1].height = 35

    # Freeze primera fila y columna CUIL
    ws.freeze_panes = 'B2'

    # Agregar hoja resumen
    ws_res = wb.create_sheet('Resumen')
    ws_res['A1'] = 'Resumen de Nómina'
    ws_res['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')

    ws_res['A3'] = 'Período:'
    ws_res['B3'] = periodo
    ws_res['A4'] = 'Municipio:'
    ws_res['B4'] = 'SAN FERNANDO'
    ws_res['A5'] = 'Total agentes:'
    ws_res['B5'] = f'=COUNTA(Datos!A2:A{ws.max_row})'
    ws_res['A6'] = 'Remun. Total bruta:'
    ws_res['B6'] = f"=SUM(Datos!J2:J{ws.max_row})"
    ws_res['B6'].number_format = '#,##0.00'
    ws_res['A7'] = 'Remun. Imponible total:'
    ws_res['B7'] = f"=SUM(Datos!K2:K{ws.max_row})"
    ws_res['B7'].number_format = '#,##0.00'
    ws_res['A8'] = 'Aporte jubilación total:'
    ws_res['B8'] = f"=SUM(Datos!S2:S{ws.max_row})"
    ws_res['B8'].number_format = '#,##0.00'

    label_font = Font(name='Arial', bold=True, size=10)
    for row in range(3, 9):
        ws_res.cell(row=row, column=1).font = label_font
        ws_res.column_dimensions['A'].width = 28
        ws_res.column_dimensions['B'].width = 20

    wb.save(output_path)
    print(f"✓ Archivo generado: {output_path}")
    print(f"  Filas de datos: {len(df)}")
    print(f"  Columnas: {len(df.columns)}")
    print(f"  Período: {periodo}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python csv_to_excel.py <archivo.csv> [salida.xlsx]")
        sys.exit(1)

    csv_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    main(csv_path, output_path)
